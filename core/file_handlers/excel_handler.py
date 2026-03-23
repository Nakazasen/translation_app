"""
Excel file handler for translation
"""
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl_image_loader import SheetImageLoader
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import io
import os
import shutil
import zipfile
import tempfile
from typing import Optional, List, Dict, Any

from translation_app.core.translator import TranslationService
from translation_app.core.ocr_handler import get_ocr_handler
from translation_app.utils.error_handler import FileProcessingError
from translation_app.utils.logger import logger


class ExcelHandler:
    """Handler for Excel file translation"""
    
    def __init__(self, translation_service: TranslationService):
        """
        Initialize Excel handler
        
        Args:
            translation_service: Translation service instance
        """
        self.translation_service = translation_service
        self.ocr_handler = get_ocr_handler()
        self.font = Font(name='Times New Roman')
    
    def translate(self, input_file: str, output_file: str, src_lang: str, dest_lang: str) -> None:
        """
        Translate Excel file while preserving images
        
        Automatically selects handler:
        - ExcelComHandler if file has drawings/external links (preserves images better)
        - openpyxl handler if file is simple (faster)
        
        Args:
            input_file: Path to input Excel file
            output_file: Path to output Excel file
            src_lang: Source language code
            dest_lang: Destination language code
        
        Raises:
            FileProcessingError: If processing fails
        """
        try:
            logger.info(f"Starting Excel translation: {input_file}")
            
            # Check if file has drawings or external links
            # If yes, use COM handler for better image preservation
            if self._has_drawings_or_external_links(input_file):
                logger.info("File contains drawings or external links. Using COM handler for better image preservation.")
                try:
                    from translation_app.core.file_handlers.excel_com_handler import ExcelComHandler, COM_AVAILABLE
                    
                    if not COM_AVAILABLE:
                        logger.warning("Excel COM not available (not Windows or pywin32 missing). Using openpyxl handler.")
                        # Fall through to openpyxl handler
                    else:
                        try:
                            com_handler = ExcelComHandler(self.translation_service)
                            com_handler.translate(input_file, output_file, src_lang, dest_lang)
                            return
                        except FileProcessingError as com_error:
                            # If it's a known error (password, locked, etc.), re-raise
                            error_str = str(com_error).lower()
                            if "password" in error_str or "locked" in error_str or "not available" in error_str:
                                raise  # Re-raise to show user the specific error
                            else:
                                logger.warning(f"COM handler failed: {com_error}. Falling back to openpyxl handler.")
                                # Fall through to openpyxl handler
                        except Exception as com_error:
                            logger.warning(f"COM handler failed with unexpected error: {com_error}. Falling back to openpyxl handler.")
                            # Fall through to openpyxl handler
                except ImportError:
                    logger.warning("Excel COM handler not available. Using openpyxl handler.")
                    # Fall through to openpyxl handler
            
            # Load workbook
            wb = load_workbook(input_file, keep_vba=False)
            
            # Store images from original file BEFORE any processing
            # Backup image data as bytes and anchor info
            original_wb = load_workbook(input_file, keep_vba=False)
            images_backup = {}
            for sheet_name in wb.sheetnames:
                if sheet_name in original_wb.sheetnames:
                    source_sheet = original_wb[sheet_name]
                    if hasattr(source_sheet, '_images') and source_sheet._images:
                        sheet_images_backup = []
                        for img in source_sheet._images:
                            try:
                                # Get image data as bytes
                                img_data = None
                                if hasattr(img, '_data'):
                                    img_data = img._data()
                                elif hasattr(img, 'ref') and hasattr(original_wb, '_rels'):
                                    # Try to get from relationships
                                    try:
                                        rel = original_wb._rels.get(img.ref)
                                        if rel and hasattr(rel, 'Target'):
                                            # This is complex, try direct copy
                                            pass
                                    except:
                                        pass
                                
                                # Get anchor info
                                anchor_info = None
                                if hasattr(img, 'anchor'):
                                    anchor = img.anchor
                                    if hasattr(anchor, '_from'):
                                        anchor_info = {
                                            'row': anchor._from.row,
                                            'col': anchor._from.col,
                                            'rowOff': getattr(anchor._from, 'rowOff', 0),
                                            'colOff': getattr(anchor._from, 'colOff', 0),
                                        }
                                    elif hasattr(anchor, 'row'):
                                        anchor_info = {
                                            'row': anchor.row,
                                            'col': anchor.col,
                                            'rowOff': 0,
                                            'colOff': 0,
                                        }
                                
                                if img_data:
                                    sheet_images_backup.append({
                                        'data': img_data,
                                        'anchor': anchor_info
                                    })
                            except Exception as e:
                                logger.warning(f"Error backing up image from sheet '{sheet_name}': {e}")
                        
                        if sheet_images_backup:
                            images_backup[sheet_name] = sheet_images_backup
                            logger.info(f"Backed up {len(sheet_images_backup)} images from sheet '{sheet_name}'")
            original_wb.close()
            
            # Count images before processing
            total_images_before = sum(len(imgs) for imgs in images_backup.values())
            logger.info(f"Found {total_images_before} images in workbook before processing")
            
            # Process all sheets
            for sheet in wb.worksheets:
                logger.info(f"Processing sheet: {sheet.title}")
                
                # Ensure images are present from backup
                if sheet.title in images_backup:
                    self._restore_sheet_images_from_backup(sheet, images_backup[sheet.title])
                
                # Collect all translation tasks
                cell_tasks = {}
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            task = self.translation_service.executor.submit(
                                self.translation_service.translate_long_text,
                                str(cell.value),
                                src_lang,
                                dest_lang
                            )
                            cell_tasks[task] = cell
                
                # Process results with timeout
                for task in cell_tasks:
                    cell = cell_tasks[task]
                    try:
                        translated_text = task.result(timeout=self.translation_service.timeout)
                        cell.value = translated_text
                        cell.font = self.font
                    except Exception as exc:
                        logger.warning(f"Error translating cell {cell.coordinate}: {exc}, keeping original")
                        # Keep original value on error
                        cell.font = self.font
                
                # OCR and translate images in sheet (this doesn't remove images, just adds text)
                if self.ocr_handler.is_installed():
                    self._process_images_in_sheet(sheet, src_lang, dest_lang)
                else:
                    logger.info("Skipping image OCR: Tesseract OCR not installed")
                
                # Re-ensure images are present after processing (in case they were lost)
                if sheet.title in images_backup:
                    current_count = len(sheet._images) if hasattr(sheet, '_images') else 0
                    backup_count = len(images_backup[sheet.title])
                    if current_count < backup_count:
                        logger.warning(f"Images lost in sheet '{sheet.title}': {current_count} < {backup_count}, restoring...")
                        self._restore_sheet_images_from_backup(sheet, images_backup[sheet.title])
            
            # Final restore of all images directly from original file before saving
            # This is more reliable than using backup
            logger.info("Restoring all images from original file before saving...")
            self._restore_all_images_direct(wb, input_file)
            
            # Save workbook
            wb.save(output_file)
            
            # Verify images in saved file
            verify_wb = load_workbook(output_file, keep_vba=False)
            total_images_saved = sum(len(sheet._images) if hasattr(sheet, '_images') else 0 
                                    for sheet in verify_wb.worksheets)
            verify_wb.close()
            
            logger.info(f"Excel translation completed: {output_file}")
            logger.info(f"Images preserved: {total_images_before} -> {total_images_saved}")
            
            if total_images_saved < total_images_before:
                logger.warning(f"Warning: Some images may not have been preserved correctly ({total_images_before} -> {total_images_saved})")
        
        except Exception as e:
            error_msg = f"Error translating Excel file: {e}"
            logger.error(error_msg, exc_info=True)
            raise FileProcessingError(error_msg, original_error=e) from e
    
    def _process_images_in_sheet(self, sheet, src_lang: str, dest_lang: str) -> None:
        """
        Process images in Excel sheet with OCR and translation
        
        Args:
            sheet: Excel worksheet
            src_lang: Source language code
            dest_lang: Destination language code
        """
        try:
            logger.info(f"Processing images in sheet '{sheet.title}'...")
            image_loader = SheetImageLoader(sheet)
            ocr_lang = self.ocr_handler.get_ocr_language(src_lang)
            images_processed = 0
            images_skipped = 0
            
            for img in sheet._images:
                try:
                    anchor = img.anchor._from if hasattr(img.anchor, '_from') else None
                    if anchor:
                        img_pil = image_loader.get(anchor.row, anchor.col)
                        if img_pil:
                            try:
                                text = self.ocr_handler.extract_text_from_image(img_pil, lang=ocr_lang)
                            except Exception as ocr_err:
                                logger.warning(f"OCR failed for image at ({anchor.row}, {anchor.col}): {ocr_err}")
                                text = ""
                            
                            # Only translate if text is clear
                            if self.ocr_handler.is_text_clear(text):
                                translated = self.translation_service.translate_long_text(text, src_lang, dest_lang)
                                row = anchor.row + 1  # Row below image
                                col = anchor.col
                                cell_ref = sheet.cell(row=row + 1, column=col + 1)
                                cell_ref.value = f"[Dịch ảnh]: {translated}"
                                cell_ref.font = self.font
                                images_processed += 1
                            else:
                                images_skipped += 1
                                logger.debug(f"Skipping image at ({anchor.row}, {anchor.col}): text not clear")
                except Exception as exc:
                    logger.warning(f"Error processing image in Excel: {exc}")
                    images_skipped += 1
            
            if images_processed > 0:
                logger.info(f"Processed {images_processed} images with clear text")
            if images_skipped > 0:
                logger.info(f"Skipped {images_skipped} images without clear text")
        
        except Exception as exc:
            logger.error(f"Error processing images in Excel sheet: {exc}")
    
    def _restore_sheet_images(self, sheet, original_file: str, sheet_name: str) -> None:
        """
        Restore images for a specific sheet from original file
        
        Args:
            sheet: Target sheet to restore images to
            original_file: Path to original Excel file
            sheet_name: Name of the sheet
        """
        try:
            original_wb = load_workbook(original_file, keep_vba=False)
            if sheet_name not in original_wb.sheetnames:
                original_wb.close()
                return
            
            source_sheet = original_wb[sheet_name]
            
            # Copy images directly from source
            if hasattr(source_sheet, '_images') and source_sheet._images:
                # Clear and restore
                sheet._images = []
                for img in source_sheet._images:
                    try:
                        # Get image data
                        if hasattr(img, '_data'):
                            img_data = img._data()
                            new_img = OpenpyxlImage(img_data)
                            
                            # Copy anchor if available
                            if hasattr(img, 'anchor'):
                                new_img.anchor = img.anchor
                            
                            sheet.add_image(new_img)
                        else:
                            # Try to copy image object directly
                            sheet._images.append(img)
                    except Exception as e:
                        logger.warning(f"Error copying image in sheet '{sheet_name}': {e}")
                
                logger.info(f"Restored {len(sheet._images)} images to sheet '{sheet_name}'")
            
            original_wb.close()
        
        except Exception as e:
            logger.error(f"Error restoring images for sheet '{sheet_name}': {e}", exc_info=True)
    
    def _restore_sheet_images_from_backup(self, sheet, images_backup: List) -> None:
        """
        Restore images to sheet from backup list
        
        Args:
            sheet: Target sheet
            images_backup: List of image backup dicts with 'data' (bytes) and 'anchor' (dict)
        """
        try:
            if not images_backup:
                return
            
            # Clear existing images
            sheet._images = []
            
            # Restore images from backup
            for img_backup in images_backup:
                try:
                    img_data = img_backup.get('data')
                    anchor_info = img_backup.get('anchor')
                    
                    if img_data:
                        # Create new image from bytes
                        new_img = OpenpyxlImage(io.BytesIO(img_data))
                        
                        # Set anchor if available
                        if anchor_info:
                            from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
                            from openpyxl.utils import get_column_letter
                            
                            try:
                                # Create anchor marker
                                from_marker = AnchorMarker(
                                    col=anchor_info['col'],
                                    colOff=anchor_info.get('colOff', 0),
                                    row=anchor_info['row'],
                                    rowOff=anchor_info.get('rowOff', 0)
                                )
                                
                                # Create two-cell anchor (simplified - assumes image size)
                                # For now, use a simple approach
                                new_img.anchor = from_marker
                            except Exception as anchor_err:
                                logger.debug(f"Could not set anchor for image: {anchor_err}, using default")
                        
                        sheet.add_image(new_img)
                except Exception as e:
                    logger.warning(f"Error restoring image to sheet '{sheet.title}': {e}")
            
            logger.info(f"Restored {len(sheet._images)} images to sheet '{sheet.title}'")
        
        except Exception as e:
            logger.error(f"Error restoring images from backup: {e}", exc_info=True)
    
    def _restore_all_images_from_backup(self, wb, images_backup: Dict[str, List]) -> None:
        """
        Restore all images from backup to workbook
        
        Args:
            wb: Workbook to restore images to
            images_backup: Dictionary mapping sheet names to image lists
        """
        try:
            for sheet_name, images_list in images_backup.items():
                if sheet_name not in wb.sheetnames:
                    continue
                
                target_sheet = wb[sheet_name]
                self._restore_sheet_images_from_backup(target_sheet, images_list)
            
            logger.info("All images restored from backup")
        
        except Exception as e:
            logger.error(f"Error restoring all images from backup: {e}", exc_info=True)
    
    def _restore_all_images_direct(self, wb, original_file: str) -> None:
        """
        Restore all images directly from original file to workbook
        This method keeps the original workbook open and copies images directly
        
        Args:
            wb: Workbook to restore images to
            original_file: Path to original Excel file
        """
        try:
            # Load original workbook - keep it open to access image data
            original_wb = load_workbook(original_file, keep_vba=False)
            
            images_restored = 0
            for sheet_name in wb.sheetnames:
                if sheet_name not in original_wb.sheetnames:
                    continue
                
                target_sheet = wb[sheet_name]
                source_sheet = original_wb[sheet_name]
                
                # Clear existing images in target
                target_sheet._images = []
                
                # Copy images from source to target
                if hasattr(source_sheet, '_images') and source_sheet._images:
                    for img in source_sheet._images:
                        try:
                            img_data = None
                            
                            # Try multiple methods to get image data
                            # Method 1: _data() method (most common)
                            if hasattr(img, '_data'):
                                try:
                                    img_data = img._data()
                                    if not isinstance(img_data, bytes):
                                        # If it's a file-like object, read it
                                        if hasattr(img_data, 'read'):
                                            img_data = img_data.read()
                                        elif hasattr(img_data, 'getvalue'):
                                            img_data = img_data.getvalue()
                                except Exception as e:
                                    logger.debug(f"_data() method failed: {e}")
                            
                            # Method 2: Try to get from workbook archive using ref
                            if not img_data and hasattr(img, 'ref'):
                                img_data = self._get_image_from_ref(original_file, img.ref, original_wb)
                            
                            # Create and add image
                            if img_data and isinstance(img_data, bytes):
                                new_img = OpenpyxlImage(io.BytesIO(img_data))
                                
                                # Copy anchor if available
                                if hasattr(img, 'anchor'):
                                    try:
                                        # Try to copy anchor directly
                                        new_img.anchor = img.anchor
                                    except Exception:
                                        # If direct copy fails, try to reconstruct
                                        try:
                                            from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
                                            anchor = img.anchor
                                            if hasattr(anchor, '_from'):
                                                from_marker = AnchorMarker(
                                                    col=anchor._from.col,
                                                    colOff=getattr(anchor._from, 'colOff', 0),
                                                    row=anchor._from.row,
                                                    rowOff=getattr(anchor._from, 'rowOff', 0)
                                                )
                                                new_img.anchor = from_marker
                                        except Exception:
                                            pass
                                
                                target_sheet.add_image(new_img)
                                images_restored += 1
                            else:
                                logger.warning(f"Could not get image data for image in sheet '{sheet_name}'")
                        except Exception as e:
                            logger.warning(f"Error restoring image in sheet '{sheet_name}': {e}")
                    
                    logger.info(f"Restored {len(target_sheet._images)} images to sheet '{sheet_name}'")
            
            original_wb.close()
            logger.info(f"Total images restored: {images_restored}")
        
        except Exception as e:
            logger.error(f"Error restoring all images directly: {e}", exc_info=True)
    
    def _get_image_from_ref(self, excel_file: str, img_ref: str, wb) -> Optional[bytes]:
        """
        Get image data from Excel file using image reference
        
        Args:
            excel_file: Path to Excel file
            img_ref: Image reference ID
            wb: Workbook object (for relationships)
            
        Returns:
            Image data as bytes, or None
        """
        try:
            # Try to get relationship to find image file
            if hasattr(wb, '_rels') and img_ref in wb._rels:
                rel = wb._rels[img_ref]
                if hasattr(rel, 'Target'):
                    # Extract image from zip
                    img_path = rel.Target.replace('../', '')
                    if not img_path.startswith('xl/'):
                        img_path = f'xl/{img_path}'
                    
                    with zipfile.ZipFile(excel_file, 'r') as zip_ref:
                        if img_path in zip_ref.namelist():
                            return zip_ref.read(img_path)
            
            return None
        
        except Exception as e:
            logger.debug(f"Error getting image from ref: {e}")
            return None
    
    def _restore_all_images(self, wb, original_file: str) -> None:
        """
        Restore all images from original file to workbook (fallback method)
        
        Args:
            wb: Workbook to restore images to
            original_file: Path to original Excel file
        """
        self._restore_all_images_direct(wb, original_file)
    
    def _has_drawings_or_external_links(self, file_path: str) -> bool:
        """
        Check if Excel file has drawings or external links
        
        Excel files are zip archives. Drawings are stored in xl/drawings/
        External links are stored in xl/externalLinks/
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            True if file has drawings or external links, False otherwise
        """
        try:
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                # Check for drawings
                drawing_files = [name for name in zip_ref.namelist() 
                               if name.startswith('xl/drawings/') and name.endswith('.xml')]
                
                # Check for external links
                external_link_files = [name for name in zip_ref.namelist() 
                                     if name.startswith('xl/externalLinks/')]
                
                has_drawings = len(drawing_files) > 0
                has_external_links = len(external_link_files) > 0
                
                if has_drawings:
                    logger.debug(f"Found {len(drawing_files)} drawing files in Excel")
                if has_external_links:
                    logger.debug(f"Found {len(external_link_files)} external link files in Excel")
                
                return has_drawings or has_external_links
        
        except Exception as e:
            logger.warning(f"Error checking for drawings/external links: {e}. Assuming no drawings.")
            return False

