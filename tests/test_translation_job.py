import concurrent.futures
import json
from pathlib import Path

from openpyxl import Workbook

from translation_app.core.file_handlers.excel_handler import ExcelHandler
from translation_app.core.translation_job import (
    TranslationJobManager,
    redact_sensitive,
)


SENTINEL_KEY = "AIzaSySentinelTestKey1234567890"


def _read_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        return []
    rows = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.strip():
            rows.append(json.loads(line))
    return rows


def test_create_job_writes_job_files(tmp_path):
    manager = TranslationJobManager(tmp_path)

    job = manager.create_job(
        input_files=["input.xlsx"],
        output_dir=tmp_path / "out",
        source_lang="en",
        target_lang="vi",
        strategy="ai",
        job_type="excel",
    )

    job_dir = tmp_path / job["job_id"]
    assert job_dir.exists()
    assert (job_dir / "job.json").exists()
    assert (job_dir / "progress.json").exists()
    assert (job_dir / "errors.jsonl").exists()
    assert (job_dir / "checkpoints.jsonl").exists()
    assert job["status"] == "pending"


def test_update_progress_updates_counts_and_percent(tmp_path):
    manager = TranslationJobManager(tmp_path)
    job = manager.create_job(["input.xlsx"], tmp_path / "out", "en", "vi", "ai")

    progress = manager.update_progress(
        job["job_id"],
        total_segments=10,
        completed_delta=3,
        failed_delta=2,
        tm_hit_delta=1,
        provider_call_delta=4,
    )

    assert progress["total_segments"] == 10
    assert progress["completed_segments"] == 3
    assert progress["failed_segments"] == 2
    assert progress["tm_hits"] == 1
    assert progress["provider_calls"] == 4
    assert progress["percent"] == 50.0


def test_record_failed_item_writes_jsonl(tmp_path):
    manager = TranslationJobManager(tmp_path)
    job = manager.create_job(["input.xlsx"], tmp_path / "out", "en", "vi", "ai")

    manager.record_failed_item(
        job["job_id"],
        file="input.xlsx",
        sheet="Sheet1",
        cell="A1",
        segment_id="Sheet1!A1",
        source_lang="en",
        target_lang="vi",
        error_type="TimeoutError",
        error_message="translation failed",
        retry_count=2,
        source_hash="abc123",
        source_length=42,
    )

    failed_items = manager.load_failed_items(job["job_id"])
    assert len(failed_items) == 1
    assert failed_items[0]["cell"] == "A1"
    assert failed_items[0]["retry_count"] == 2
    assert failed_items[0]["source_hash"] == "abc123"


def test_checkpoint_append_jsonl(tmp_path):
    manager = TranslationJobManager(tmp_path)
    job = manager.create_job(["input.xlsx"], tmp_path / "out", "en", "vi", "ai")

    manager.record_checkpoint(job["job_id"], "segment_started", file="input.xlsx", sheet="Sheet1", cell="A1", segment_id="Sheet1!A1", status="running")
    manager.record_checkpoint(job["job_id"], "segment_completed", file="input.xlsx", sheet="Sheet1", cell="A1", segment_id="Sheet1!A1", status="completed")

    checkpoints = _read_jsonl(tmp_path / job["job_id"] / "checkpoints.jsonl")
    assert len(checkpoints) == 2
    assert checkpoints[0]["event"] == "segment_started"
    assert checkpoints[1]["event"] == "segment_completed"


def test_can_resume_statuses(tmp_path):
    manager = TranslationJobManager(tmp_path)
    resumable = {"pending", "running", "paused", "failed"}
    non_resumable = {"completed", "cancelled"}

    for status in resumable | non_resumable:
        job = manager.create_job([f"{status}.xlsx"], tmp_path / "out", "en", "vi", "ai")
        manager.update_job_status(job["job_id"], status)
        if status in resumable:
            assert manager.can_resume(job["job_id"]) is True
        else:
            assert manager.can_resume(job["job_id"]) is False


def test_job_files_do_not_store_api_keys(tmp_path):
    manager = TranslationJobManager(tmp_path)
    job = manager.create_job(
        input_files=[f"input_{SENTINEL_KEY}.xlsx"],
        output_dir=tmp_path / "out",
        source_lang="en",
        target_lang="vi",
        strategy=f"ai:{SENTINEL_KEY}",
        notes=f"notes {SENTINEL_KEY}",
    )

    manager.record_failed_item(
        job["job_id"],
        error_type="RuntimeError",
        error_message=f"Provider rejected key {SENTINEL_KEY}",
    )
    manager.record_checkpoint(
        job["job_id"],
        "segment_started",
        file=f"in_{SENTINEL_KEY}.xlsx",
        sheet="Sheet1",
        cell="A1",
        segment_id="Sheet1!A1",
        status="running",
    )

    job_dir = tmp_path / job["job_id"]
    for filename in ("job.json", "progress.json", "errors.jsonl", "checkpoints.jsonl"):
        contents = (job_dir / filename).read_text(encoding="utf-8")
        assert SENTINEL_KEY not in contents
        assert "[REDACTED_API_KEY]" in contents or filename == "progress.json"


def test_retry_failed_items_preserves_retry_count(tmp_path):
    manager = TranslationJobManager(tmp_path)
    job = manager.create_job(["input.xlsx"], tmp_path / "out", "en", "vi", "ai")
    manager.record_failed_item(job["job_id"], cell="A1", segment_id="Sheet1!A1", error_message="fail once", retry_count=1)

    retry_items = manager.prepare_retry_failed(job["job_id"])

    assert len(retry_items) == 1
    assert retry_items[0]["retry_count"] == 2
    assert manager.load_failed_items(job["job_id"])[0]["retry_count"] == 1


def test_atomic_write_does_not_corrupt_existing_job(tmp_path, monkeypatch):
    manager = TranslationJobManager(tmp_path)
    path = tmp_path / "job.json"
    original = {"status": "old"}
    path.write_text(json.dumps(original), encoding="utf-8")

    def fail_replace(src, dst):
        raise OSError("replace failed")

    monkeypatch.setattr("translation_app.core.translation_job.os.replace", fail_replace)

    try:
        manager._write_json_atomic(path, {"status": "new"})
    except OSError:
        pass

    assert json.loads(path.read_text(encoding="utf-8")) == original
    assert not path.with_suffix(".json.tmp").exists()


def test_excel_translation_job_hook_if_feasible(tmp_path, monkeypatch):
    workbook_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Hello"
    ws["B2"] = "World"
    wb.save(workbook_path)

    manager = TranslationJobManager(tmp_path / "jobs")

    class FakeOCRHandler:
        def is_installed(self):
            return False

    class FakeTranslationService:
        def __init__(self):
            self.executor = concurrent.futures.ThreadPoolExecutor(max_workers=2)
            self.timeout = 5
            self.strategy = "ai"
            self.observer = None

        def set_runtime_observer(self, observer):
            self.observer = observer

        def clear_runtime_observer(self):
            self.observer = None

        def translate_long_text(self, text, src_lang, dest_lang):
            if self.observer:
                self.observer("provider_call", {"provider": "fake"})
            return f"{text}-{dest_lang}"

    service = FakeTranslationService()
    handler = ExcelHandler(service)
    handler.ocr_handler = FakeOCRHandler()

    monkeypatch.setattr("translation_app.core.file_handlers.excel_handler.get_translation_job_manager", lambda: manager)
    monkeypatch.setattr(ExcelHandler, "_has_drawings_or_external_links", lambda self, _: False)
    monkeypatch.setattr(ExcelHandler, "_restore_all_images_direct", lambda self, wb_obj, input_file: None)

    try:
        handler.translate(str(workbook_path), str(output_path), "en", "vi")
    finally:
        service.executor.shutdown(wait=True)

    jobs = manager.list_jobs(limit=1)
    assert len(jobs) == 1
    summary = manager.get_job_summary(jobs[0]["job_id"])
    assert summary["job"]["status"] == "completed"
    assert summary["progress"]["total_segments"] == 2
    assert summary["progress"]["completed_segments"] == 2
    assert summary["progress"]["provider_calls"] == 2


def test_job_redacts_api_key_like_patterns():
    redacted = redact_sensitive(f"bad key {SENTINEL_KEY}")
    assert SENTINEL_KEY not in redacted
    assert "[REDACTED_API_KEY]" in redacted

