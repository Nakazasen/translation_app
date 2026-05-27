import concurrent.futures
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

from translation_app.core.file_handlers.excel_handler import ExcelHandler
from translation_app.core.translation_job import TranslationJobManager


class FakeOCRHandler:
    def is_installed(self):
        return False


class FakeTranslationService:
    def __init__(self, translations=None):
        self.executor = concurrent.futures.ThreadPoolExecutor(max_workers=4)
        self.timeout = 5
        self.strategy = "ai"
        self.observer = None
        self.calls = []
        self.translations = translations or {}

    def set_runtime_observer(self, observer):
        self.observer = observer

    def clear_runtime_observer(self):
        self.observer = None

    def translate_long_text(self, text, src_lang, dest_lang):
        self.calls.append(text)
        if self.observer:
            self.observer("provider_call", {"provider": "fake"})
        return self.translations.get(text, f"{text}-{dest_lang}")


def _run_excel_translation(tmp_path, monkeypatch, build_workbook, translations=None):
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    manager = TranslationJobManager(tmp_path / "jobs")

    workbook = Workbook()
    build_workbook(workbook)
    workbook.save(input_path)
    workbook.close()

    service = FakeTranslationService(translations=translations)
    handler = ExcelHandler(service)
    handler.ocr_handler = FakeOCRHandler()

    monkeypatch.setattr(
        "translation_app.core.file_handlers.excel_handler.get_translation_job_manager",
        lambda: manager,
    )
    monkeypatch.setattr(ExcelHandler, "_has_drawings_or_external_links", lambda self, _: False)

    try:
        handler.translate(str(input_path), str(output_path), "en", "vi")
    finally:
        service.executor.shutdown(wait=True)

    return output_path, service, manager


def _latest_job_summary(manager: TranslationJobManager):
    jobs = manager.list_jobs(limit=1)
    assert len(jobs) == 1
    return manager.get_job_summary(jobs[0]["job_id"])


def test_excel_skips_formula_cells(tmp_path, monkeypatch):
    def build_workbook(workbook):
        ws = workbook.active
        ws["A1"] = "Hello"
        ws["B1"] = "=SUM(1,2)"

    output_path, service, manager = _run_excel_translation(
        tmp_path,
        monkeypatch,
        build_workbook,
        translations={"Hello": "Xin chao"},
    )

    result = load_workbook(output_path)
    ws = result.active
    assert ws["A1"].value == "Xin chao"
    assert ws["B1"].value == "=SUM(1,2)"
    assert service.calls == ["Hello"]

    summary = _latest_job_summary(manager)
    assert summary["progress"]["total_segments"] == 1
    assert summary["progress"]["completed_segments"] == 1
    assert summary["progress"]["provider_calls"] == 1
    result.close()


def test_excel_preserves_cell_styles(tmp_path, monkeypatch):
    def build_workbook(workbook):
        ws = workbook.active
        cell = ws["A1"]
        cell.value = "Styled text"
        cell.font = Font(name="Calibri", size=14, bold=True, italic=True, color="00FF00")
        cell.fill = PatternFill(fill_type="solid", fgColor="FFCC00")
        cell.border = Border(left=Side(style="thin", color="0000FF"))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.number_format = "@"
        cell.protection = Protection(locked=False, hidden=False)

    output_path, _, _ = _run_excel_translation(tmp_path, monkeypatch, build_workbook)

    result = load_workbook(output_path)
    cell = result.active["A1"]
    assert cell.value == "Styled text-vi"
    assert cell.font.name == "Calibri"
    assert cell.font.bold is True
    assert cell.font.italic is True
    assert cell.fill.fill_type == "solid"
    assert cell.fill.fgColor.rgb == "00FFCC00"
    assert cell.border.left.style == "thin"
    assert cell.alignment.horizontal == "center"
    assert cell.alignment.vertical == "center"
    assert cell.alignment.wrap_text is True
    assert cell.number_format == "@"
    assert cell.protection.locked is False
    result.close()


def test_excel_preserves_merged_cells(tmp_path, monkeypatch):
    def build_workbook(workbook):
        ws = workbook.active
        ws.merge_cells("A1:C1")
        ws["A1"] = "Merged title"

    output_path, service, _ = _run_excel_translation(tmp_path, monkeypatch, build_workbook)

    result = load_workbook(output_path)
    ws = result.active
    assert "A1:C1" in [str(cell_range) for cell_range in ws.merged_cells.ranges]
    assert ws["A1"].value == "Merged title-vi"
    assert ws["B1"].value is None
    assert ws["C1"].value is None
    assert service.calls == ["Merged title"]
    result.close()


def test_excel_preserves_row_column_dimensions(tmp_path, monkeypatch):
    def build_workbook(workbook):
        ws = workbook.active
        ws["A1"] = "Size check"
        ws.column_dimensions["A"].width = 22.5
        ws.column_dimensions["A"].hidden = True
        ws.row_dimensions[1].height = 35
        ws.row_dimensions[1].hidden = True

    output_path, _, _ = _run_excel_translation(tmp_path, monkeypatch, build_workbook)

    result = load_workbook(output_path)
    ws = result.active
    assert ws["A1"].value == "Size check-vi"
    assert ws.column_dimensions["A"].width == pytest.approx(22.5, rel=0, abs=0.1)
    assert ws.column_dimensions["A"].hidden is True
    assert ws.row_dimensions[1].height == pytest.approx(35, rel=0, abs=0.1)
    assert ws.row_dimensions[1].hidden is True
    result.close()


def test_excel_preserves_comments(tmp_path, monkeypatch):
    def build_workbook(workbook):
        ws = workbook.active
        ws["A1"] = "Annotated"
        ws["A1"].comment = Comment("Keep this note", "tester")

    output_path, _, _ = _run_excel_translation(tmp_path, monkeypatch, build_workbook)

    result = load_workbook(output_path)
    comment = result.active["A1"].comment
    assert result.active["A1"].value == "Annotated-vi"
    assert comment is not None
    assert comment.text == "Keep this note"
    assert comment.author == "tester"
    result.close()


def test_excel_preserves_sheet_names_and_order(tmp_path, monkeypatch):
    def build_workbook(workbook):
        workbook.active.title = "Summary"
        workbook.active["A1"] = "Overview"
        details = workbook.create_sheet("Details")
        details["B2"] = "Line item"
        archive = workbook.create_sheet("Archive")
        archive["C3"] = "Past"

    output_path, service, _ = _run_excel_translation(tmp_path, monkeypatch, build_workbook)

    result = load_workbook(output_path)
    assert result.sheetnames == ["Summary", "Details", "Archive"]
    assert result["Summary"]["A1"].value == "Overview-vi"
    assert result["Details"]["B2"].value == "Line item-vi"
    assert result["Archive"]["C3"].value == "Past-vi"
    assert service.calls == ["Overview", "Line item", "Past"]
    result.close()


def test_excel_does_not_translate_numbers_dates_booleans(tmp_path, monkeypatch):
    date_value = datetime(2024, 1, 2, 3, 4, 5)

    def build_workbook(workbook):
        ws = workbook.active
        ws["A1"] = "Translate me"
        ws["B1"] = 123.45
        ws["C1"] = date_value
        ws["D1"] = True

    output_path, service, _ = _run_excel_translation(tmp_path, monkeypatch, build_workbook)

    result = load_workbook(output_path)
    ws = result.active
    assert ws["A1"].value == "Translate me-vi"
    assert ws["B1"].value == 123.45
    assert ws["C1"].value == date_value
    assert ws["D1"].value is True
    assert service.calls == ["Translate me"]
    result.close()


def test_excel_job_segment_count_excludes_formulas_and_non_text(tmp_path, monkeypatch):
    def build_workbook(workbook):
        ws = workbook.active
        ws["A1"] = "Hello"
        ws["B1"] = "=SUM(1,2)"
        ws["C1"] = 42
        ws["D1"] = False
        ws["E1"] = "World"

    _, service, manager = _run_excel_translation(tmp_path, monkeypatch, build_workbook)

    summary = _latest_job_summary(manager)
    assert service.calls == ["Hello", "World"]
    assert summary["progress"]["total_segments"] == 2
    assert summary["progress"]["completed_segments"] == 2
    assert summary["progress"]["provider_calls"] == 2


def test_excel_preserves_images_if_supported(tmp_path, monkeypatch):
    try:
        from PIL import Image as PILImage
    except ImportError:
        pytest.skip("Pillow not installed")

    image_path = tmp_path / "tiny.png"
    PILImage.new("RGB", (8, 8), color="red").save(image_path)

    def build_workbook(workbook):
        ws = workbook.active
        ws["A1"] = "Caption"
        ws.add_image(OpenpyxlImage(str(image_path)), "C3")

    output_path, service, _ = _run_excel_translation(tmp_path, monkeypatch, build_workbook)

    result = load_workbook(output_path)
    ws = result.active
    assert ws["A1"].value == "Caption-vi"
    assert service.calls == ["Caption"]
    assert len(getattr(ws, "_images", [])) == 1
    result.close()
